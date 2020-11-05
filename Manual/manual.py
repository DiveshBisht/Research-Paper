import openpyxl

w = openpyxl.load_workbook("manual_label_final.xlsx")
sheet = w.get_sheet_by_name("Sheet1")

for i in range (2, sheet.max_row + 1):
    pos = 0
    neg = 0
    for j in range (5,9):
        if(sheet.cell(row = i, column = j).value == "love"):
            pos = pos + 1
        if(sheet.cell(row = i, column = j).value == "joy"):
            pos = pos + 1
        if(sheet.cell(row = i, column = j).value == "fear"):
            neg = neg + 1
        if(sheet.cell(row = i, column = j).value == "anger"):
            neg = neg + 1
        if(sheet.cell(row = i, column = j).value == "sadness"):
            neg = neg + 1
    summ = pos + neg
    if pos>=3 and neg==0:
        result = 'positive'
    elif neg>=3 and pos==0:
        result = 'negative'
    elif summ==0 or summ==1:
        result = 'neutral'
    else:
        result = 'contradictory'    
    sheet.cell(row = i, column = 9).value = result

w.save('manual_mapping_final.xlsx')

