import openpyxl

w = openpyxl.load_workbook("Ratings.xlsx")
sheet = w.get_sheet_by_name("Sheet1")

col1 = sheet.cell(row = 1, column = 8).value
col2 = sheet.cell(row = 1, column = 1).value
v1,v2,v3,v4,v5,v6,v7,v8,v9 = 0,0,0,0,0,0,0,0,0
for i in range (2, sheet.max_row + 1):
    x = sheet.cell(row = i, column = 8).value
    y = sheet.cell(row = i, column = 1).value   
    if(x == "neutral" and y == "neutral"):
        v1 = v1 + 1
    if(x == "neutral" and y == "positive"):
        v2 = v2 + 1
    if(x == "neutral" and y == "negative"):
        v3 = v3 + 1
    if(x == "positive" and y == "neutral"):
        v4 = v4 + 1
    if(x == "positive" and y == "positive"):
        v5 = v5 + 1
    if(x == "positive" and y == "negative"):
        v6 = v6 + 1
    if(x == "negative" and y == "neutral"):
        v7 = v7 + 1
    if(x == "negative" and y == "positive"):
        v8 = v8 + 1
    if(x == "negative" and y == "negative"):
        v9 = v9 + 1
print(col1 + "vs" + col2)
print(v1,v2,v3)
print(v4,v5,v6)
print(v7,v8,v9)  


