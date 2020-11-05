import openpyxl


w = openpyxl.load_workbook("Ratings.xlsx")
sheet = w.get_sheet_by_name("Sheet1")

for i in range(2, sheet.max_row + 1):
    if(sheet.cell(row = i, column = 2).value == "neutral" and sheet.cell(row = i, column = 3).value == "neutral" and sheet.cell(row = i, column = 4).value == "neutral"):
        result = "neutral"
    
    elif(sheet.cell(row = i, column = 2).value == "positive" and sheet.cell(row = i, column = 3).value == "positive" and sheet.cell(row = i, column = 4).value == "positive"):
                result = "positive"
    
    elif(sheet.cell(row = i, column = 2).value == "negative" and sheet.cell(row = i, column = 3).value == "negative" and sheet.cell(row = i, column = 4).value == "negative"):
                result = "negative"
    else:
        result = "invalid"
    sheet.cell(row = i, column = 8).value = result

w.save("Int.xlsx")
    
