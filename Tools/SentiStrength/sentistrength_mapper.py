import openpyxl

w = openpyxl.load_workbook("ss_data.xlsx")
sheet = w.get_sheet_by_name("Sheet1")

for i in range (2, sheet.max_row + 1):
	pos = sheet.cell(row = i, column = 7).value
	neg = sheet.cell(row = i, column = 8).value	
	var = pos + neg
	if(var > 0):
		sheet.cell(row = i, column = 10).value = 1
	if(var < 0):
		sheet.cell(row = i, column = 10).value = -1
	if(var == 0 and pos<4):
		sheet.cell(row = i, column = 10).value = 0
	if(var == 0 and pos>=4):
		sheet.cell(row = i, column = 10).value = 100

w.save('ss_result.xlsx')
