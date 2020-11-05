import openpyxl


w = openpyxl.load_workbook("RQ1.xlsx")
sheet = w.get_sheet_by_name("Sheet1")

for i in range(2, sheet.max_row + 1):
    if(sheet.cell(row = i, column = 2).value == "love" and sheet.cell(row = i, column = 3).value == "love" and sheet.cell(row = i, column = 4).value == "love" and sheet.cell(row = i, column = 5).value == "love"):
        result = "love"
    
    elif(sheet.cell(row = i, column = 2).value == "joy" and sheet.cell(row = i, column = 3).value == "joy" and sheet.cell(row = i, column = 4).value == "joy" and sheet.cell(row = i, column = 5).value == "joy"):
        result = "joy"

    elif(sheet.cell(row = i, column = 2).value == "surprise" and sheet.cell(row = i, column = 3).value == "surprise" and sheet.cell(row = i, column = 4).value == "surprise" and sheet.cell(row = i, column = 5).value == "surprise"):
        result = "surprise"
    
    elif(sheet.cell(row = i, column = 2).value == "anger" and sheet.cell(row = i, column = 3).value == "anger" and sheet.cell(row = i, column = 4).value == "anger" and sheet.cell(row = i, column = 5).value == "anger"):
        result = "anger"

    elif(sheet.cell(row = i, column = 2).value == "sadness" and sheet.cell(row = i, column = 3).value == "sadness" and sheet.cell(row = i, column = 4).value == "sadness" and sheet.cell(row = i, column = 5).value == "sadness"):
        result = "sadness"

    elif(sheet.cell(row = i, column = 2).value == "fear" and sheet.cell(row = i, column = 3).value == "fear" and sheet.cell(row = i, column = 4).value == "fear" and sheet.cell(row = i, column = 5).value == "fear"):
        result = "fear"

    elif(sheet.cell(row = i, column = 2).value == "nil" and sheet.cell(row = i, column = 3).value == "nil" and sheet.cell(row = i, column = 4).value == "nil" and sheet.cell(row = i, column = 5).value == "nil"):
        result = "nil"
    else:
        result = "invalid"
    sheet.cell(row = i, column = 6).value = result

w.save("Int.xlsx")
    
