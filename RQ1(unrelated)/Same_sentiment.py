import openpyxl

w = openpyxl.load_workbook("RQ1.xlsx")
sheet = w.get_sheet_by_name("Sheet1")

for i in range (2, sheet.max_row + 1):
    pos = 0
    neg = 0
    for j in range (2,6):
        if(sheet.cell(row = i, column = j).value == "love"):
            pos = pos + 1
        if(sheet.cell(row = i, column = j).value == "joy"):
            pos = pos + 1
        if(sheet.cell(row = i, column = j).value == "fear"):
            neg = neg + 1
        if(sheet.cell(row = i, column = j).value == "anger"):
            neg = neg + 1
        if(sheet.cell(row = i, column = j).value == "sadness"):
            neg = neg + 1
    summ = pos + neg
    if pos>=4 and neg==0:
        result = 'positive'
    elif neg>=4 and pos==0:
        result = 'negative'
    elif summ==0:
        result = 'neutral'
    else:
        result = 'invalid'    
    sheet.cell(row = i, column = 9).value = result

w.save('manual_RQ1.xlsx')
